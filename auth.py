from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from urllib.parse import urlparse
from models import db, User
from forms import LoginForm, RegistrationForm, AdminApprovalForm, UserSearchForm
from datetime import datetime
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from werkzeug.utils import secure_filename


auth = Blueprint('auth', __name__)

@auth.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password', 'error')
            return redirect(url_for('auth.login'))
        
        if not user.can_login():
            if not user.is_approved and not user.is_admin():
                flash('Your account is pending approval. Please wait for admin approval.', 'warning')
            elif not user.is_active:
                flash('Your account has been suspended. Please contact admin.', 'error')
            return redirect(url_for('auth.login'))
        
        login_user(user, remember=form.remember_me.data)
        next_page = request.args.get('next')
        if not next_page or urlparse(next_page).netloc != '':
            next_page = url_for('home')
        
        flash(f'Welcome back, {user.full_name}!', 'success')
        return redirect(next_page)
    
    return render_template('auth/login.html', title='Login', form=form, page_title='Login - Daiva Anughara')

@auth.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        # Check for existing username
        existing_user = User.query.filter_by(username=form.username.data).first()
        if existing_user:
            flash('Username already taken. Please choose a different one.', 'error')
            return render_template('auth/register.html', title='Register', form=form)
        
        # Check for existing email
        existing_email = User.query.filter_by(email=form.email.data).first()
        if existing_email:
            flash('Email already registered. Please use a different one.', 'error')
            return render_template('auth/register.html', title='Register', form=form)
        
        # Handle profile picture upload
        profile_picture_path = None
        if form.profile_picture.data:
            file = form.profile_picture.data
            if file and file.filename:
                # Create uploads directory if it doesn't exist
                upload_dir = os.path.join(os.getcwd(), 'static', 'uploads', 'profiles')
                os.makedirs(upload_dir, exist_ok=True)
                
                # Generate secure filename
                filename = secure_filename(file.filename)
                # Add timestamp to make filename unique
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                name, ext = os.path.splitext(filename)
                filename = f"{name}_{timestamp}{ext}"
                
                # Save file
                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)
                profile_picture_path = f"uploads/profiles/{filename}"
        
        user = User(
            username=form.username.data,
            email=form.email.data,
            full_name=form.full_name.data,
            phone=form.phone.data,
            spiritual_name=form.spiritual_name.data,
            guru_name=form.guru_name.data,
            practice_level=form.practice_level.data,
            purpose=form.purpose.data,
            profile_picture=profile_picture_path
        )
        user.set_password(form.password.data)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please wait for admin approval before you can login.', 'success')
        return redirect(url_for('auth.login'))
    
    return render_template('auth/register.html', title='Register', form=form, page_title='Register - Daiva Anughara')

@auth.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('home'))

@auth.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))
    
    search_form = UserSearchForm()
    approval_form = AdminApprovalForm()
    
    # Get search parameters
    search_term = request.args.get('search_term', '')
    status_filter = request.args.get('status_filter', 'all')
    role_filter = request.args.get('role_filter', 'all')
    
    # Build query
    query = User.query
    
    if search_term:
        query = query.filter(
            db.or_(
                User.username.contains(search_term),
                User.email.contains(search_term),
                User.full_name.contains(search_term)
            )
        )
    
    if status_filter != 'all':
        if status_filter == 'pending':
            query = query.filter_by(is_approved=False, is_active=True)
        elif status_filter == 'approved':
            query = query.filter_by(is_approved=True, is_active=True)
        elif status_filter == 'rejected':
            query = query.filter_by(is_approved=False, is_active=False)
        elif status_filter == 'suspended':
            query = query.filter_by(is_active=False)
    
    if role_filter != 'all':
        query = query.filter_by(role=role_filter)
    
    users = query.order_by(User.created_at.desc()).all()
    
    return render_template('admin/users.html', 
                         title='User Management',
                         page_title='User Management - Daiva Anughara',
                         users=users,
                         search_form=search_form,
                         approval_form=approval_form)

@auth.route('/admin/approve_user', methods=['POST'])
@login_required
def approve_user():
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    form = AdminApprovalForm()
    if form.validate_on_submit():
        user_id = form.user_id.data
        action = form.action.data
        notes = form.notes.data
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        if action == 'approve':
            user.is_approved = True
            user.approved_at = datetime.utcnow()
            user.approved_by = current_user.id
            user.is_active = True
            message = f'User {user.username} has been approved successfully.'
        elif action == 'reject':
            user.is_approved = False
            user.is_active = False
            message = f'User {user.username} has been rejected.'
        elif action == 'suspend':
            user.is_active = False
            message = f'User {user.username} has been suspended.'
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': message})
    
    return jsonify({'success': False, 'message': 'Invalid form data'}), 400

@auth.route('/admin/user/<int:user_id>')
@login_required
def admin_user_detail(user_id):
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))
    
    user = User.query.get_or_404(user_id)
    return render_template('admin/user_detail.html', title='User Detail', page_title='User Detail - Daiva Anughara', user=user)

@auth.route('/admin/user/<int:user_id>/mandala-access', methods=['POST'])
@login_required
def update_mandala_access(user_id):
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('auth.admin_users'))
    
    user = User.query.get_or_404(user_id)
    
    # Get mandala access updates from form
    mandala_2_access = request.form.get('mandala_2_access') == 'on'
    mandala_3_access = request.form.get('mandala_3_access') == 'on'
    
    # Update mandala access
    user.mandala_2_access = mandala_2_access
    user.mandala_3_access = mandala_3_access
    
    db.session.commit()
    
    flash(f'Mandala access updated for {user.username}', 'success')
    return redirect(url_for('auth.admin_user_detail', user_id=user_id))

@auth.route('/profile')
@login_required
def profile():
    return render_template('auth/profile.html', title='Profile', page_title='Profile - Daiva Anughara')

@auth.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profile():
    # This would be implemented for users to edit their own profile
    pass

def generate_user_report():
    users = User.query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "User Report"

    # Headers
    headers = [
        "User ID", "Username", "Full Name", "Email", "Role", "Status",
        "Join Date", "Approval Date", "Approved for Mandala 2", "Approved for Mandala 3",
        "Days on Mandala 1", "Days on Mandala 2", "Days on Mandala 3"
    ]
    sheet.append(headers)

    # Style for headers
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for user in users:
        status = "Admin" if user.is_admin() else ("Approved" if user.is_approved else ("Suspended" if not user.is_active else "Pending"))
        
        days_on_mandala_1 = (datetime.utcnow() - user.created_at).days if user.created_at else 0
        days_on_mandala_2 = 0
        days_on_mandala_3 = 0

        if user.mandala_2_access and user.approved_at:
            days_on_mandala_2 = (datetime.utcnow() - user.approved_at).days
        
        if user.mandala_3_access and user.approved_at: # Assuming mandala 3 access is granted at the same time or after mandala 2
            days_on_mandala_3 = (datetime.utcnow() - user.approved_at).days


        sheet.append([
            user.id,
            user.username,
            user.full_name,
            user.email,
            user.role,
            status,
            user.created_at.strftime("%Y-%m-%d") if user.created_at else "",
            user.approved_at.strftime("%Y-%m-%d") if user.approved_at else "",
            "Yes" if user.mandala_2_access else "No",
            "Yes" if user.mandala_3_access else "No",
            days_on_mandala_1,
            days_on_mandala_2,
            days_on_mandala_3
        ])

    # KPIs
    sheet.cell(row=1, column=15, value="KPIs").font = header_font
    
    total_users = len(users)
    approved_users = len([u for u in users if u.is_approved])
    pending_users = len([u for u in users if not u.is_approved and u.is_active])
    
    sheet.cell(row=2, column=15, value="Total Users")
    sheet.cell(row=2, column=16, value=total_users)
    sheet.cell(row=3, column=15, value="Approved Users")
    sheet.cell(row=3, column=16, value=approved_users)
    sheet.cell(row=4, column=15, value="Pending Users")
    sheet.cell(row=4, column=16, value=pending_users)

    # Chart
    chart_sheet = workbook.create_sheet(title="User Status Chart")
    chart_data = [
        ['Status', 'Count'],
        ['Approved', approved_users],
        ['Pending', pending_users],
        ['Suspended', total_users - approved_users - pending_users]
    ]
    for row in chart_data:
        chart_sheet.append(row)
    
    chart = BarChart()
    chart.title = "User Status Distribution"
    chart.y_axis.title = "Number of Users"
    chart.x_axis.title = "Status"
    
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=4, max_col=2)
    cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart_sheet.add_chart(chart, "E5")


    # Save to a BytesIO object
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file

@auth.route('/admin/users/report')
@login_required
def user_report():
    if not current_user.is_admin():
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('home'))

    excel_file = generate_user_report()

    return send_file(
        excel_file,
        as_attachment=True,
        download_name='user_report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@auth.route('/admin/users/kpi_data')
@login_required
def user_kpi_data():
    if not current_user.is_admin():
        return jsonify({'error': 'Access denied'}), 403

    users = User.query.all()

    total_users = len(users)
    approved_users = len([u for u in users if u.is_approved and u.is_active])
    pending_users = len([u for u in users if not u.is_approved and u.is_active])
    suspended_users = len([u for u in users if not u.is_active and not u.is_admin()])

    mandala_2_users = len([u for u in users if u.mandala_2_access])
    mandala_3_users = len([u for u in users if u.mandala_3_access])

    # Calculate average days to approval
    approved_users_with_approval_date = [u for u in users if u.is_approved and u.approved_at and u.created_at]
    if approved_users_with_approval_date:
        total_days_to_approval = sum([(u.approved_at - u.created_at).days for u in approved_users_with_approval_date])
        average_days_to_approval = total_days_to_approval / len(approved_users_with_approval_date)
    else:
        average_days_to_approval = 0

    data = {
        'kpis': {
            'total_users': total_users,
            'approved_users': approved_users,
            'pending_users': pending_users,
            'suspended_users': suspended_users,
            'average_days_to_approval': round(average_days_to_approval, 2)
        },
        'charts': {
            'user_status_distribution': {
                'labels': ['Approved', 'Pending', 'Suspended'],
                'data': [approved_users, pending_users, suspended_users]
            },
            'mandala_access_distribution': {
                'labels': ['Mandala 2 Access', 'Mandala 3 Access'],
                'data': [mandala_2_users, mandala_3_users]
            }
        }
    }
    return jsonify(data)